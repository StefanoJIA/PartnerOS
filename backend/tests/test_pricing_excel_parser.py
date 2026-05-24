"""Unit tests for D6.2.1 pricing Excel parser."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest

from app.services.quotes.pricing_excel_parser import (
    COST_ALIASES,
    classify_sheet,
    detect_header_row,
    generate_internal_sku,
    normalize_header,
    parse_cost_sheet,
    parse_decimal_cell,
    parse_percent_cell,
    parse_quantity_range,
    parse_workbook_sheet,
)


def test_normalize_header_chinese_english_symbols():
    assert normalize_header("  成本RMB \n") == "成本rmb"
    assert normalize_header("FOB成本（美金）") == "fob成本 美金"
    assert normalize_header("Product Name") == "product name"


def test_detect_header_row_not_first_row():
    rows = [
        ("", "", "海运单价", "汇率"),
        ("", "", "22", "6.7"),
        ("", "", "", ""),
        ("", "Product Name", "成本RMB", "重量", "FOB成本美金"),
        ("恒升", "Desk Frame", "100", "30", "150"),
    ]
    det = detect_header_row(rows, COST_ALIASES, min_score=3)
    assert det is not None
    assert det.row_index == 3
    assert "product" in det.columns
    assert "material_cost" in det.columns


def test_classify_sheet_by_name():
    assert classify_sheet("成本表", []) == "cost_model"
    assert classify_sheet("价目表", []) == "price_tier"
    assert classify_sheet("Quote HOSUN", []) == "quote_template"
    assert classify_sheet("利润计算器", []) == "calculator"


def test_classify_sheet_by_header_aliases():
    rows = [
        ("products", "MinQty", "MaxQty", "FOB", "DDP"),
        ("Desk", "1", "49", "100", "120"),
    ]
    assert classify_sheet("Sheet1", rows) == "price_tier"


@pytest.mark.parametrize(
    "value,expected",
    [
        ("1-49", {"min_qty": 1, "max_qty": 49}),
        ("50–99", {"min_qty": 50, "max_qty": 99}),
        ("500+", {"min_qty": 500, "max_qty": None}),
        ("≥500", {"min_qty": 500, "max_qty": None}),
        ("500以上", {"min_qty": 500, "max_qty": None}),
        ("1 ~ 49", {"min_qty": 1, "max_qty": 49}),
    ],
)
def test_parse_quantity_range(value, expected):
    assert parse_quantity_range(value) == expected


@pytest.mark.parametrize(
    "value,expected",
    [
        ("$123.45", Decimal("123.45")),
        ("¥123.45", Decimal("123.45")),
        ("1,234.56", Decimal("1234.56")),
        ("RMB 123.45", Decimal("123.45")),
        ("-", None),
        (None, None),
    ],
)
def test_parse_decimal_cell(value, expected):
    assert parse_decimal_cell(value) == expected


@pytest.mark.parametrize(
    "value,expected",
    [
        ("10%", Decimal("0.10")),
        ("0.1", Decimal("0.1")),
        ("10", Decimal("0.10")),
    ],
)
def test_parse_percent_cell(value, expected):
    assert parse_percent_cell(value) == expected


def test_generate_sku_from_product_name():
    used: set[str] = set()
    sku = generate_internal_sku("Adjustable Desk Frame", partner_code="HOSUN", used=used)
    assert sku.startswith("HOSUN-")
    sku2 = generate_internal_sku("Adjustable Desk Frame", partner_code="HOSUN", used=used)
    assert sku2.endswith("-002")


def test_generate_sku_deduplication():
    used: set[str] = set()
    a = generate_internal_sku("DF0102", partner_code="JOOBOO", explicit_sku="DF0102", used=used)
    b = generate_internal_sku("DF0102", partner_code="JOOBOO", explicit_sku="DF0102", used=used)
    assert a == "DF0102"
    assert b == "DF0102-002"


def test_parse_cost_sheet_sample_rows():
    rows = [
        ("", "", "", "海运单价", "汇率", "国内利润%"),
        ("", "", "", "22", "6.7", "9"),
        ("", "", "", "", "", ""),
        ("", "Product Name", "成本RMB", "重量", "运输成本", "运费成本美金", "FOB成本美金", "DDP成本美金"),
        ("恒升", "Desk Frame A", "1000", "30", "500", "80", "150", "230"),
    ]
    report = parse_cost_sheet("成本", rows)
    assert report.header_row == 4
    assert report.candidate_rows == 1
    assert report.cost_rows[0].partner_code == "HOSUN"
    assert report.meta.fx_rate == Decimal("6.7")


def test_import_handles_missing_sheet_gracefully():
    report = parse_workbook_sheet("Unknown", [("a", "b"), ("1", "2")])
    assert report.detected_type == "unknown"


@pytest.mark.skipif(
    not Path(__file__).resolve().parents[2].joinpath("local_data", "报价模型与格式.xlsx").is_file(),
    reason="local Excel workbook not present",
)
def test_real_workbook_nonzero_candidates():
    from openpyxl import load_workbook

    from app.services.quotes.pricing_excel_parser import load_sheet_rows

    xlsx = Path(__file__).resolve().parents[2] / "local_data" / "报价模型与格式.xlsx"
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    total = 0
    for name in wb.sheetnames:
        report = parse_workbook_sheet(name, load_sheet_rows(wb[name]))
        total += report.candidate_rows
    wb.close()
    assert total > 0
