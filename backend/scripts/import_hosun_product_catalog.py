"""Import HOSUN product catalog rows and assign IntelliOpus SKUs.

This script imports HOSUN product master data only. It does not create customer
quotes, does not create orders, and does not expose costs or margins.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import load_workbook

from app.models import ManufacturingPartner, ProductCatalog


DEFAULT_PRODUCT_CATALOG = Path(r"E:\WORKS\07 2025 Fall Winter\IntelliOpus\Herstar\HOSUN 2026 Product Catalog.xlsx")
DEFAULT_CLASSIFICATION = Path(
    r"E:\WORKS\08 2026 Spring\IntelliOpus\IntelliOffice\HOSUN\产品分类.20260226.xlsx"
)


def cell_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", cell_text(value).lower()).strip("_")


def normalize_model(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", cell_text(value).upper())


def category_prefix(category: str, name: str) -> str:
    text = f"{category} {name}".lower()
    if "heavy" in text or "1320" in text:
        return "HD"
    if "pneumatic" in text:
        return "PD"
    if "column" in text:
        return "LC"
    if "bench" in text or "face-to-face" in text or "combination" in text or "workstation" in text:
        return "BF"
    if "accessor" in text or "control" in text or "handset" in text:
        return "AC"
    if "frame" in text or "desk" in text:
        return "DF"
    return "LS"


def generate_hosun_sku(model: str, category: str, name: str) -> str:
    clean_model = normalize_model(model)
    prefix = category_prefix(category, name)
    if clean_model:
        return f"IO-HOSUN-{prefix}-{clean_model}"[:64]
    slug = re.sub(r"[^A-Z0-9]+", "-", name.upper()).strip("-")[:36]
    return f"IO-HOSUN-{prefix}-{slug}"[:64]


def taxonomy(category: str, name: str) -> tuple[str, str]:
    text = f"{category} {name}".lower()
    if "pneumatic" in text:
        return "lifting_systems", "pneumatic_standing_desks"
    if "heavy" in text:
        return "lifting_systems", "heavy_duty_supply"
    if "column" in text:
        return "lifting_systems", "lifting_columns"
    if "bench" in text or "face-to-face" in text or "combination" in text or "workstation" in text:
        return "lifting_systems", "benching_frames"
    if "frame" in text or "desk" in text:
        return "lifting_systems", "desk_frames"
    if "accessor" in text or "control" in text:
        return "lifting_systems", "desk_accessories"
    return "lifting_systems", "hosun_general"


def infer_image_url(category: str, name: str, model: str) -> str | None:
    text = f"{category} {name} {model}".lower()
    if "pneumatic" in text:
        if "v-leg" in text:
            return "/desk-order-assets/products/V-LEG.png"
        if "easylift" in text:
            return "/desk-order-assets/products/EASYLIFT.png"
        return "/desk-order-assets/products/pneumatic-desks.png"
    if "bench" in text or "face-to-face" in text or "workstation" in text:
        return "/desk-order-assets/products/multi-user-benching.png"
    if "column" in text:
        if "90603" in text:
            return "/desk-order-assets/products/90X60正装三节立柱-Photoroom.png"
        if "90602" in text:
            return "/desk-order-assets/products/90X60正装两节立柱-Photoroom.png"
        if "80503" in text:
            return "/desk-order-assets/products/80X50正装三节立柱-Photoroom.png"
        if "80502" in text:
            return "/desk-order-assets/products/80X50正装两节立柱-Photoroom.png"
        if "70703" in text:
            return "/desk-order-assets/products/70X70正装三节立柱-Photoroom.png"
        if "70702" in text:
            return "/desk-order-assets/products/70X70正装两节立柱-Photoroom.png"
        if "round" in text or "007" in text:
            return "/desk-order-assets/products/圆形正装三节立柱-Photoroom.png"
        return "/desk-order-assets/products/electric-columns.png"
    if "accessor" in text:
        return "/desk-order-assets/products/accessories.png"
    if "heavy" in text:
        return "/desk-order-assets/products/standalone-frames.png"
    if "frame" in text or "desk" in text:
        if "single" in text:
            return "/desk-order-assets/products/80x50单电机桌架-Photoroom.png"
        return "/desk-order-assets/products/standalone-frames.png"
    return None


def load_product_catalog(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [normalize_key(cell.value) for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        current_category = ""
        for raw in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[idx]: raw[idx] if idx < len(raw) else None for idx in range(len(headers))}
            category = cell_text(data.get("category")) or current_category
            current_category = category or current_category
            model = normalize_model(data.get("model"))
            name = cell_text(data.get("product_name"))
            if not model or not name:
                continue
            rows.append(
                {
                    "category": category,
                    "model": model,
                    "product_name": name,
                    "tube_size": cell_text(data.get("tube_size")),
                    "height_range": cell_text(data.get("height_range")),
                    "width_range": cell_text(data.get("width_range")),
                    "load_capacity": cell_text(data.get("load_capacity")),
                    "speed": cell_text(data.get("speed")),
                    "source_sheet": ws.title,
                }
            )
        return rows
    finally:
        wb.close()


def load_classification(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.is_file() or path.name.startswith("~$"):
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out: dict[str, dict[str, Any]] = {}
        current_group = ""
        for raw in ws.iter_rows(min_row=4, values_only=True):
            group = cell_text(raw[0]) or current_group
            current_group = group or current_group
            model = normalize_model(raw[2] if len(raw) > 2 else None)
            if not model:
                continue
            out[model] = {
                "classification_group": group,
                "specification": cell_text(raw[3] if len(raw) > 3 else None),
                "classification_name": cell_text(raw[4] if len(raw) > 4 else None),
                "chinese_name": cell_text(raw[5] if len(raw) > 5 else None),
                "stages": cell_text(raw[6] if len(raw) > 6 else None),
                "product_class": cell_text(raw[7] if len(raw) > 7 else None),
                "lifting_range": cell_text(raw[8] if len(raw) > 8 else None),
                "adjustable_width": cell_text(raw[9] if len(raw) > 9 else None),
                "dynamic_load": cell_text(raw[10] if len(raw) > 10 else None),
                "lifting_speed": cell_text(raw[11] if len(raw) > 11 else None),
            }
        return out
    finally:
        wb.close()


def ensure_partner(db) -> ManufacturingPartner:
    partner = db.query(ManufacturingPartner).filter(ManufacturingPartner.partner_code == "HOSUN").first()
    if partner:
        return partner
    partner = ManufacturingPartner(
        partner_name="HOSUN",
        partner_type="Lifting System Manufacturer",
        country="China",
        partner_code="HOSUN",
        default_incoterm="FOB",
        default_currency="USD",
        catalog_status="active",
        notes="HOSUN product catalog import.",
    )
    db.add(partner)
    db.flush()
    return partner


def unique_sku(db, desired: str, *, current_product_id=None) -> str:
    candidate = desired[:64]
    suffix = 2
    while True:
        hit = db.query(ProductCatalog).filter(ProductCatalog.internal_sku == candidate).first()
        if not hit or hit.id == current_product_id:
            return candidate
        base = desired[:58]
        candidate = f"{base}-{suffix:02d}"[:64]
        suffix += 1


def find_product(db, partner: ManufacturingPartner, *, model: str, sku: str, name: str) -> ProductCatalog | None:
    hit = (
        db.query(ProductCatalog)
        .filter(ProductCatalog.partner_id == partner.id, ProductCatalog.partner_product_code == model)
        .first()
    )
    if hit:
        return hit
    hit = db.query(ProductCatalog).filter(ProductCatalog.internal_sku == sku).first()
    if hit:
        return hit
    return (
        db.query(ProductCatalog)
        .filter(ProductCatalog.partner_id == partner.id, ProductCatalog.product_name.ilike(name.strip()))
        .first()
    )


def run(
    *,
    product_catalog: Path,
    classification: Path | None,
    apply: bool,
    overwrite: bool,
    overwrite_sku: bool,
) -> int:
    from app.core.database import SessionLocal

    if not product_catalog.is_file():
        print(f"HOSUN product catalog not found: {product_catalog}")
        return 1
    rows = load_product_catalog(product_catalog)
    class_map = load_classification(classification)
    summary = {"created": 0, "updated": 0, "skipped": 0, "sku_updated": 0, "images_mapped": 0}
    db = SessionLocal()
    try:
        partner = ensure_partner(db)
        for row in rows:
            model = row["model"]
            name = row["product_name"]
            category = row["category"] or class_map.get(model, {}).get("classification_group") or "HOSUN"
            sku = generate_hosun_sku(model, category, name)
            category_code, family = taxonomy(category, name)
            image_url = infer_image_url(category, name, model)
            existing = find_product(db, partner, model=model, sku=sku, name=name)
            attrs = dict(existing.attributes_json if existing and existing.attributes_json else {})
            attrs.update(
                {
                    "source_system": "hosun_product_catalog",
                    "source_workbook": product_catalog.name,
                    "classification_workbook": classification.name if classification else None,
                    "partner_model": model,
                    "intelliopus_sku_rule": "IO-HOSUN-{category-prefix}-{model}",
                    "tube_size": row.get("tube_size"),
                    "height_range": row.get("height_range"),
                    "width_range": row.get("width_range"),
                    "load_capacity": row.get("load_capacity"),
                    "lifting_speed": row.get("speed"),
                    "hosun_classification": class_map.get(model),
                    "product_family_hint": family,
                    "product_category_hint": category_code,
                    "customer_safe_pricing_mode": "full_quantity_interval_quote_table",
                }
            )
            if existing:
                if not overwrite:
                    summary["skipped"] += 1
                    continue
                if overwrite_sku:
                    new_sku = unique_sku(db, sku, current_product_id=existing.id)
                    if existing.internal_sku != new_sku:
                        existing.internal_sku = new_sku
                        summary["sku_updated"] += 1
                existing.partner_product_code = model
                existing.product_name = name
                existing.product_category = category_code
                existing.product_family = family
                existing.description_customer = name
                existing.attributes_json = attrs
                if image_url:
                    existing.image_url = image_url
                    summary["images_mapped"] += 1
                existing.notes = "hosun_product_catalog_import"
                summary["updated"] += 1
                continue
            product = ProductCatalog(
                partner_id=partner.id,
                internal_sku=unique_sku(db, sku),
                partner_product_code=model,
                product_name=name,
                product_category=category_code,
                product_family=family,
                description_customer=name,
                status="active",
                default_uom="EA",
                base_currency="USD",
                default_incoterm="FOB",
                image_url=image_url,
                attributes_json=attrs,
                notes="hosun_product_catalog_import",
            )
            db.add(product)
            db.flush()
            summary["created"] += 1
            if image_url:
                summary["images_mapped"] += 1
        if apply:
            db.commit()
            print("HOSUN product catalog import applied.")
        else:
            db.rollback()
            print("Dry-run only; no database changes.")
    finally:
        db.close()
    print("HOSUN Product Catalog Import Summary")
    print(f"  source rows: {len(rows)}")
    for key, value in summary.items():
        print(f"  {key}: {value}")
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Import HOSUN product catalog and assign IntelliOpus SKUs")
    parser.add_argument("--product-catalog", default=str(DEFAULT_PRODUCT_CATALOG))
    parser.add_argument("--classification", default=str(DEFAULT_CLASSIFICATION))
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--overwrite-sku", action="store_true")
    args = parser.parse_args()
    if args.apply and not args.confirm:
        print("Refusing --apply without --confirm")
        sys.exit(1)
    sys.exit(
        run(
            product_catalog=Path(args.product_catalog),
            classification=Path(args.classification) if args.classification else None,
            apply=args.apply and args.confirm,
            overwrite=args.overwrite,
            overwrite_sku=args.overwrite_sku,
        )
    )


if __name__ == "__main__":
    main()
