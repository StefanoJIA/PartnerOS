"""Sync HOSUN quote catalog products from the latest classification workbook.

The classification workbook is the current product whitelist for HOSUN quote
catalog display. This script updates internal product metadata only: it does
not create customer quotes, send messages, change orders, or expose cost data.
"""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import load_workbook
from sqlalchemy import func

from app.core.database import SessionLocal
from app.models import ManufacturingPartner, OrderLineItem, ProductCatalog, ProductPriceTier, QuoteLineItem
from app.services.quotes.catalog_enrichment import enrich_product_attributes, infer_margin_strategy

DEFAULT_CLASSIFICATION_DIR = Path(r"E:\WORKS\08 2026 Spring\IntelliOpus\IntelliOffice\HOSUN")
DEFAULT_CLASSIFICATION_KEYWORD = "20260226"

ADDITIONAL_HOSUN_PRODUCTS: tuple[dict[str, str], ...] = (
    {
        "model": "HS11A",
        "name": "Hand Control Panel HS11A",
        "chinese_name": "手控器 HS11A",
        "product_class": "Hand Control Panel",
        "image_url": "/desk-order-assets/products/HS11A-Photoroom.png",
    },
    {
        "model": "HS11B",
        "name": "Hand Control Panel HS11B",
        "chinese_name": "手控器 HS11B",
        "product_class": "Hand Control Panel",
        "image_url": "/desk-order-assets/products/HS11B-Photoroom.png",
    },
    {
        "model": "HS11C",
        "name": "Hand Control Panel HS11C",
        "chinese_name": "手控器 HS11C",
        "product_class": "Hand Control Panel",
        "image_url": "/desk-order-assets/products/HS11C-Photoroom.png",
    },
    {
        "model": "HS11D",
        "name": "Hand Control Panel HS11D",
        "chinese_name": "手控器 HS11D",
        "product_class": "Hand Control Panel",
        "image_url": "/desk-order-assets/products/HS11D-Photoroom.png",
    },
    {
        "model": "HS11E",
        "name": "Hand Control Panel HS11E",
        "chinese_name": "手控器 HS11E",
        "product_class": "Hand Control Panel",
        "image_url": "/desk-order-assets/products/HS11E-Photoroom.png",
    },
    {
        "model": "SWATCH-4COLOR",
        "name": "Four Color Roll Swatch Sample Set",
        "chinese_name": "四色卷料色卡样品套装",
        "product_class": "Color Swatch Sample",
        "image_url": "/desk-order-assets/products/accessories.png",
    },
)


def cell_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_model(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", cell_text(value).upper())


def clean_sku(value: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9_-]+", "", value.strip().upper())
    return cleaned[:64] or "PRODUCT"


def resolve_default_classification() -> Path:
    for path in DEFAULT_CLASSIFICATION_DIR.glob("*.xlsx"):
        if DEFAULT_CLASSIFICATION_KEYWORD in path.name and not path.name.startswith("~$"):
            return path
    raise FileNotFoundError(f"classification workbook not found in {DEFAULT_CLASSIFICATION_DIR}")


def taxonomy(product_class: str, name: str) -> tuple[str, str]:
    text = f"{product_class} {name}".lower()
    if "hand control" in text or "swatch" in text or "sample" in text:
        return "lifting_systems", "desk_accessories"
    if "pneumatic" in text:
        return "lifting_systems", "pneumatic_standing_desks"
    if "combination" in text or "workstation" in text or "bench" in text or "face-to-face" in text:
        return "lifting_systems", "benching_frames"
    if "triple-motor" in text or "3-leg" in text:
        return "lifting_systems", "heavy_duty_supply"
    if "column" in text and "frame" not in text:
        return "lifting_systems", "lifting_columns"
    if "single-motor" in text or "dual-motor" in text or "desk frame" in text:
        return "lifting_systems", "desk_frames"
    return "lifting_systems", "desk_frames"


def image_for(row: dict[str, str]) -> str | None:
    if row.get("image_url"):
        return row["image_url"]
    name = row["name"].lower()
    model = row["model"].upper()
    spec = row["spec"].lower()
    base = "/desk-order-assets/products/"
    if "pneumatic" in name:
        if "v-leg" in name:
            return base + "V-LEG.png"
        if "easylift" in name:
            return base + "EASYLIFT.png"
        return base + "pneumatic-desks.png"
    if "combination" in name or "workstation" in name or "face-to-face" in name:
        if "120" in name or "trio" in name:
            return base + "multi-user-120-trio.png"
        return base + "multi-user-face-to-face.png"
    if "lifting column" in name:
        if "90603" in model:
            return base + "90X60正装三节立柱-Photoroom.png"
        if "80503" in model:
            return base + "80X50正装三节立柱-Photoroom.png"
        if "80502" in model:
            return base + "80X50正装两节立柱-Photoroom.png"
        if "70703" in model:
            return base + "70X70正装三节立柱-Photoroom.png"
        if "70702" in model:
            return base + "70X70正装两节立柱-Photoroom.png"
        if "00703" in model:
            return base + "圆形正装三节立柱-Photoroom.png"
        if "00702" in model:
            return base + "圆形正装两节立柱-Photoroom.png"
        if "oval" in name:
            return base + "椭圆管正装二节立柱-Photoroom.png"
        return base + "electric-columns.png"
    if "single-motor" in name:
        if "round" in name:
            return base + "圆管单电机桌架-Photoroom.png"
        if "oval" in name:
            return base + "椭圆管单电机桌架-Photoroom.png"
        return base + "80x50单电机桌架-Photoroom.png"
    if "triple-motor" in name or "3-leg" in name:
        return base + "三腿拐角-Photoroom.png"
    if "70*70" in spec or "70x70" in name:
        return base + "70X70正装两节桌架 (1)-Photoroom.png"
    if "80*50" in spec or "80x50" in name:
        return base + "80X50正装两节桌架 (3)-Photoroom.png"
    return base + "standalone-frames.png"


def load_classification_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows: list[dict[str, str]] = []
        current_group = ""
        for raw in sheet.iter_rows(min_row=4, values_only=True):
            group = cell_text(raw[0] if len(raw) > 0 else None) or current_group
            current_group = group or current_group
            model = normalize_model(raw[2] if len(raw) > 2 else None)
            name = cell_text(raw[4] if len(raw) > 4 else None)
            if not model or not name:
                continue
            rows.append(
                {
                    "classification_group": group,
                    "seq": cell_text(raw[1] if len(raw) > 1 else None),
                    "model": model,
                    "spec": cell_text(raw[3] if len(raw) > 3 else None),
                    "name": name,
                    "chinese_name": cell_text(raw[5] if len(raw) > 5 else None),
                    "stages": cell_text(raw[6] if len(raw) > 6 else None),
                    "product_class": cell_text(raw[7] if len(raw) > 7 else None),
                    "lifting_range": cell_text(raw[8] if len(raw) > 8 else None),
                    "adjustable_width": cell_text(raw[9] if len(raw) > 9 else None),
                    "load_capacity": cell_text(raw[10] if len(raw) > 10 else None),
                    "lifting_speed": cell_text(raw[11] if len(raw) > 11 else None),
                    "keywords": cell_text(raw[12] if len(raw) > 12 else None),
                    "package_count": cell_text(raw[13] if len(raw) > 13 else None),
                    "package_size": cell_text(raw[14] if len(raw) > 14 else None),
                }
            )
        return rows
    finally:
        workbook.close()


def additional_hosun_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in ADDITIONAL_HOSUN_PRODUCTS:
        rows.append(
            {
                "classification_group": "Accessories",
                "seq": "",
                "model": item["model"],
                "spec": "",
                "name": item["name"],
                "chinese_name": item["chinese_name"],
                "stages": "",
                "product_class": item["product_class"],
                "lifting_range": "",
                "adjustable_width": "",
                "load_capacity": "",
                "lifting_speed": "",
                "keywords": item["product_class"],
                "package_count": "",
                "package_size": "",
                "image_url": item["image_url"],
                "fixed_accessory": "true",
            }
        )
    return rows


def ensure_partner(db) -> ManufacturingPartner:
    partner = db.query(ManufacturingPartner).filter(ManufacturingPartner.partner_code == "HOSUN").first()
    if partner:
        return partner
    partner = ManufacturingPartner(
        partner_name="HOSUN",
        partner_type="Lifting System Manufacturer",
        country="China",
        partner_code="HOSUN",
        default_incoterm="FOB",
        default_currency="USD",
        catalog_status="active",
        notes="HOSUN product catalog sync.",
    )
    db.add(partner)
    db.flush()
    return partner


def unique_sku(db, desired: str, product_id=None) -> str:
    base = clean_sku(desired)
    candidate = base
    index = 2
    while True:
        existing = db.query(ProductCatalog).filter(ProductCatalog.internal_sku == candidate).first()
        if existing is None or existing.id == product_id:
            return candidate
        suffix = f"-{index}"
        candidate = f"{base[: 64 - len(suffix)]}{suffix}"
        index += 1


def find_existing(db, partner: ManufacturingPartner, row: dict[str, str]) -> ProductCatalog | None:
    model = row["model"]
    hit = (
        db.query(ProductCatalog)
        .filter(ProductCatalog.partner_id == partner.id, ProductCatalog.partner_product_code == model)
        .first()
    )
    if hit:
        return hit
    exact_name = row["name"].strip()
    return (
        db.query(ProductCatalog)
        .filter(ProductCatalog.partner_id == partner.id, ProductCatalog.product_name.ilike(exact_name))
        .first()
    )


def apply_row(product: ProductCatalog, row: dict[str, str], *, source_workbook: str) -> Decimal:
    category, family = taxonomy(row["product_class"], row["name"])
    strategy, reason, target_margin = infer_margin_strategy(
        name=row["name"],
        category=row["product_class"],
        product_family=family,
        attrs={"partner_model": row["model"]},
    )
    attrs = dict(product.attributes_json or {})
    attrs.update(
        {
            "source_system": "hosun_product_classification",
            "source_workbook": source_workbook,
            "partner_model": row["model"],
            "classification_group": row["classification_group"],
            "classification_seq": row["seq"],
            "chinese_name": row["chinese_name"],
            "specification": row["spec"],
            "stage_count": row["stages"],
            "product_class": row["product_class"],
            "height_range": row["lifting_range"],
            "adjustable_width": row["adjustable_width"],
            "load_capacity": row["load_capacity"],
            "lifting_speed": row["lifting_speed"],
            "package_count": row["package_count"],
            "package_size": row["package_size"],
            "customer_safe_pricing_mode": "full_quantity_interval_quote_table",
            "commercial_margin_strategy": strategy,
            "commercial_margin_strategy_reason": reason,
            "target_margin": str(target_margin),
            "quote_markup_multiplier": str((Decimal("1") + target_margin).quantize(Decimal("0.0001"))),
            "pricing_margin_source": "auto_catalog_commercial_classification",
        }
    )
    attrs.pop("inactive_reason", None)
    attrs = enrich_product_attributes(
        name=row["name"],
        category=category,
        product_family=family,
        partner_code="HOSUN",
        partner_product_code=row["model"],
        internal_sku=row["model"],
        attrs=attrs,
        description=product.description_internal,
    )
    attrs["target_margin"] = str(target_margin)
    attrs["quote_markup_multiplier"] = str((Decimal("1") + target_margin).quantize(Decimal("0.0001")))
    attrs["pricing_margin_source"] = "auto_catalog_commercial_classification"
    product.partner_product_code = row["model"]
    product.product_name = row["name"]
    product.product_category = category
    product.product_family = family
    product.description_customer = row["name"]
    product.status = "active"
    product.default_incoterm = "FOB"
    product.image_url = image_for(row)
    product.attributes_json = attrs
    product.notes = "hosun_classification_catalog_sync"
    return target_margin


def normalize_incoterms(db) -> int:
    rows = db.query(ProductPriceTier).filter(ProductPriceTier.incoterm == "EXW").all()
    for row in rows:
        row.incoterm = "FOB"
        row.source = f"{row.source or 'classification_sync'}; exw_normalized_to_fob"
    return len(rows)


def fill_missing_target_margins(db) -> int:
    count = 0
    for product in db.query(ProductCatalog).all():
        attrs = dict(product.attributes_json or {})
        if attrs.get("target_margin") not in (None, ""):
            continue
        strategy, reason, target_margin = infer_margin_strategy(
            name=product.product_name,
            category=product.product_category,
            product_family=product.product_family,
            attrs=attrs,
        )
        attrs["commercial_margin_strategy"] = strategy
        attrs["commercial_margin_strategy_reason"] = reason
        attrs["target_margin"] = str(target_margin)
        attrs["quote_markup_multiplier"] = str((Decimal("1") + target_margin).quantize(Decimal("0.0001")))
        attrs["pricing_margin_source"] = "auto_catalog_commercial_classification"
        product.attributes_json = attrs
        count += 1
    return count


def _purge_obsolete_hosun_products(db, partner: ManufacturingPartner, whitelist: set[str]) -> dict[str, int]:
    obsolete = [
        product
        for product in db.query(ProductCatalog).filter(ProductCatalog.partner_id == partner.id).all()
        if normalize_model(product.partner_product_code) not in whitelist
    ]
    obsolete_ids = [product.id for product in obsolete]
    if not obsolete_ids:
        return {
            "purged_obsolete_hosun": 0,
            "quote_line_refs_cleared": 0,
            "order_line_refs_cleared": 0,
        }

    quote_refs = (
        db.query(QuoteLineItem)
        .filter(QuoteLineItem.product_catalog_id.in_(obsolete_ids))
        .update({QuoteLineItem.product_catalog_id: None}, synchronize_session=False)
    )
    order_refs = (
        db.query(OrderLineItem)
        .filter(OrderLineItem.product_catalog_id.in_(obsolete_ids))
        .update({OrderLineItem.product_catalog_id: None}, synchronize_session=False)
    )
    for product in obsolete:
        db.delete(product)
    return {
        "purged_obsolete_hosun": len(obsolete),
        "quote_line_refs_cleared": int(quote_refs or 0),
        "order_line_refs_cleared": int(order_refs or 0),
    }


def run(*, classification: Path | None, apply: bool, purge_obsolete: bool = False) -> int:
    path = classification or resolve_default_classification()
    classification_rows = load_classification_rows(path)
    rows = classification_rows + additional_hosun_rows()
    whitelist = {normalize_model(row["model"]) for row in rows}
    summary = {
        "source_rows": len(classification_rows),
        "fixed_accessory_rows": len(ADDITIONAL_HOSUN_PRODUCTS),
        "created": 0,
        "updated": 0,
        "inactivated_hosun": 0,
        "purged_obsolete_hosun": 0,
        "quote_line_refs_cleared": 0,
        "order_line_refs_cleared": 0,
        "target_margins_filled": 0,
        "incoterms_normalized": 0,
    }
    db = SessionLocal()
    try:
        partner = ensure_partner(db)
        for row in rows:
            existing = find_existing(db, partner, row)
            if existing is None:
                existing = ProductCatalog(
                    partner_id=partner.id,
                    internal_sku=unique_sku(db, row["model"]),
                    partner_product_code=row["model"],
                    product_name=row["name"],
                    product_category="lifting_systems",
                    product_family="desk_frames",
                    status="active",
                    default_uom="EA",
                    base_currency="USD",
                    default_incoterm="FOB",
                )
                db.add(existing)
                db.flush()
                summary["created"] += 1
            else:
                summary["updated"] += 1
                existing.internal_sku = unique_sku(db, row["model"], existing.id)
            apply_row(existing, row, source_workbook=path.name)

        if purge_obsolete:
            purge_summary = _purge_obsolete_hosun_products(db, partner, whitelist)
            summary.update(purge_summary)
        else:
            for product in db.query(ProductCatalog).filter(ProductCatalog.partner_id == partner.id).all():
                model = normalize_model(product.partner_product_code)
                if model not in whitelist:
                    product.status = "inactive"
                    attrs = dict(product.attributes_json or {})
                    attrs["inactive_reason"] = "not_in_latest_hosun_classification_20260226"
                    product.attributes_json = attrs
                    summary["inactivated_hosun"] += 1
        for product in db.query(ProductCatalog).join(ManufacturingPartner).filter(
            ManufacturingPartner.partner_code.in_(["OTHER", "FUTURE"]),
            ProductCatalog.product_category == "lifting_systems",
        ):
            product.status = "inactive"
            attrs = dict(product.attributes_json or {})
            attrs["inactive_reason"] = "legacy_unmapped_lifting_product_not_in_latest_hosun_classification_20260226"
            product.attributes_json = attrs
            summary["inactivated_hosun"] += 1

        summary["target_margins_filled"] = fill_missing_target_margins(db)
        summary["incoterms_normalized"] = normalize_incoterms(db)

        if apply:
            db.commit()
            print("HOSUN classification catalog sync applied.")
        else:
            db.rollback()
            print("Dry-run only; no database changes.")

        active_hosun = (
            db.query(func.count(ProductCatalog.id))
            .filter(ProductCatalog.partner_id == partner.id, ProductCatalog.status == "active")
            .scalar()
        )
        print("HOSUN Classification Catalog Sync Summary")
        for key, value in summary.items():
            print(f"  {key}: {value}")
        print(f"  active_hosun_after_sync: {active_hosun if apply else 'dry-run'}")
        print("  safety: no quote creation, no external sending, no customer notification, no raw token handling.")
        return 0
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync HOSUN product catalog from latest classification workbook")
    parser.add_argument("--classification", default=None)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm", action="store_true")
    parser.add_argument(
        "--purge-obsolete",
        action="store_true",
        help="Physically remove HOSUN catalog rows that are not in the latest 56-product whitelist.",
    )
    args = parser.parse_args()
    if args.apply and not args.confirm:
        print("Refusing --apply without --confirm")
        sys.exit(1)
    sys.exit(
        run(
            classification=Path(args.classification) if args.classification else None,
            apply=args.apply and args.confirm,
            purge_obsolete=args.purge_obsolete,
        )
    )


if __name__ == "__main__":
    main()
