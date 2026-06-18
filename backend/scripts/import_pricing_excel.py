"""Import pricing Excel from local_data into product catalog (D6.2 / D6.2.1)."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.models import (
    FxRate,
    ManufacturingPartner,
    MarginStrategyTier,
    ProductCatalog,
    ProductCostModel,
    ProductPriceTier,
)
from app.services.quotes.pricing_excel_parser import (
    ParsedCostRow,
    ParsedMarginRow,
    ParsedPriceRow,
    SheetParseReport,
    format_sheet_debug,
    load_sheet_rows,
    parse_workbook_sheet,
    parse_decimal_cell,
    parse_quantity_range,
)

DEFAULT_PARTNERS = (
    {"code": "HOSUN", "name": "HOSUN Lifting Systems", "type": "Lifting System Manufacturer"},
    {"code": "JOOBOO", "name": "JOOBOO Education Furniture", "type": "Education Furniture Manufacturer"},
    {"code": "CHONGQING_HUIJU", "name": "Chongqing Huiju", "type": "Manufacturing Partner"},
    {"code": "OTHER", "name": "Other Partner", "type": "Manufacturing Partner"},
)

CUSTOMER_QUOTE_MODEL_SOURCE = "excel_quote_template"
INTERNAL_PRICE_MODEL_SOURCE = "excel_fixed_cost_interval_model"


def _product_taxonomy(name: str, partner_code: str | None) -> tuple[str, str]:
    text = name.lower()
    partner = (partner_code or "").upper()
    if partner == "JOOBOO" or any(token in text for token in ("classroom", "chair", "single desk", "education")):
        if "chair" in text:
            return "education_furniture", "school_chairs"
        if "desk" in text:
            return "education_furniture", "school_desks"
        return "education_furniture", "project_furniture"
    if any(token in text for token in ("pneumatic", "standing desk")):
        return "lifting_systems", "pneumatic_standing_desks"
    if any(token in text for token in ("benching", "face-to-face", "four-motor")):
        return "lifting_systems", "benching_frames"
    if any(token in text for token in ("heavy duty", "heavy-duty", "300kg")):
        return "lifting_systems", "heavy_duty_desk_frames"
    if any(token in text for token in ("desk frame", "dual-motor", "tri-motor", "lifting")):
        return "lifting_systems", "desk_frames"
    return "other", "general"


def _infer_partner_code_from_product_name(name: str, fallback: str | None = None) -> str:
    text = name.lower()
    if _partner_product_code_from_name(name):
        return "JOOBOO"
    if any(token in text for token in ("desk frame", "benching frame", "dual-motor", "tri-motor", "triple-motor", "pneumatic")):
        return "HOSUN"
    return fallback or "OTHER"


def _pricing_model_attributes(existing: dict | None, *, product_name: str, partner_code: str | None) -> dict:
    category, family = _product_taxonomy(product_name, partner_code)
    data = dict(existing or {})
    data.update(
        {
            "source_workbook": "报价模型与格式.xlsx",
            "pricing_model": "fixed_rmb_cost_plus_ocean_freight_fx_plus_interval_margin",
            "pricing_model_source": INTERNAL_PRICE_MODEL_SOURCE,
            "customer_safe_pricing_mode": "full_quantity_interval_quote_table",
            "internal_only_cost_model": True,
            "product_category_hint": category,
            "product_family_hint": family,
            "pricing_model_steps": [
                "fixed RMB product cost is maintained per product",
                "transport cost = unit weight * ocean freight unit price",
                "FOB cost USD = RMB cost / USD-CNY exchange rate",
                "DDP cost USD = (RMB cost + transport cost) / USD-CNY exchange rate",
                "customer interval price = selected cost basis * interval margin multiplier plus workbook adjustment",
                "quote output shows every quantity interval for each selected product",
            ],
        }
    )
    return data


def _infer_product_image_url(product_name: str, partner_code: str | None) -> str | None:
    """Map workbook product names to imported PartnerOS product assets when the source workbook has no image."""

    text = product_name.lower()
    partner = (partner_code or "").upper()
    if partner == "JOOBOO":
        return None
    if "3-leg" in text or "triple-motor" in text:
        return "/desk-order-assets/products/三腿拐角-Photoroom.png"
    if "face-to-face" in text or "benching" in text or "four-motor" in text:
        return "/desk-order-assets/products/multi-user-face-to-face.png"
    if "pneumatic" in text and "v-leg" in text:
        return "/desk-order-assets/products/V-LEG.png"
    if "pneumatic" in text and ("easylift" in text or "rectangular" in text):
        return "/desk-order-assets/products/EASYLIFT.png"
    if "pneumatic" in text:
        return "/desk-order-assets/products/STANDARD.png"
    if "single-motor" in text and "round" in text:
        return "/desk-order-assets/products/圆管单电机桌架-Photoroom.png"
    if "single-motor" in text:
        return "/desk-order-assets/products/80x50单电机桌架-Photoroom.png"
    if "round" in text:
        return "/desk-order-assets/products/圆形正装两节立柱-Photoroom.png"
    if "heavy" in text:
        return "/desk-order-assets/products/70X70正装两节桌架 (1)-Photoroom.png"
    if "3-stage" in text and ("3.54" in text or "90x60" in text):
        return "/desk-order-assets/products/90X60正装三节立柱-Photoroom.png"
    if "3-stage" in text:
        return "/desk-order-assets/products/80X50正装三节立柱-Photoroom.png"
    if "2-stage" in text and ("3.54" in text or "90x60" in text):
        return "/desk-order-assets/products/80X50正装两节桌架 (3)-Photoroom.png"
    if "desk frame" in text or "adjustable desk frame" in text:
        return "/desk-order-assets/products/70X70正装两节桌架 (1)-Photoroom.png"
    if "desk legs" in text:
        return "/desk-order-assets/products/standalone-frames.png"
    if "lifting column" in text:
        return "/desk-order-assets/products/electric-columns.png"
    return None


def _assign_product_metadata(product: ProductCatalog, *, product_name: str, partner_code: str | None) -> None:
    category, family = _product_taxonomy(product_name, partner_code)
    product.product_category = category
    product.product_family = family
    product.attributes_json = _pricing_model_attributes(
        product.attributes_json,
        product_name=product_name,
        partner_code=partner_code,
    )
    if not product.image_url:
        product.image_url = _infer_product_image_url(product_name, partner_code)


def _ensure_partner(db, code: str, *, overwrite: bool) -> tuple[ManufacturingPartner, str]:
    row = db.query(ManufacturingPartner).filter(ManufacturingPartner.partner_code == code).first()
    if row:
        return row, "skipped"
    spec = next((p for p in DEFAULT_PARTNERS if p["code"] == code), None)
    name = spec["name"] if spec else f"{code} Partner"
    ptype = spec["type"] if spec else "Manufacturing Partner"
    row = ManufacturingPartner(
        partner_name=name,
        partner_type=ptype,
        country="China",
        partner_code=code,
        default_incoterm="FOB",
        default_currency="USD",
        catalog_status="active",
        notes="excel_import",
    )
    db.add(row)
    db.flush()
    return row, "created"


def _find_product(db, *, sku: str | None, name: str, partner_id) -> ProductCatalog | None:
    if sku:
        hit = db.query(ProductCatalog).filter(ProductCatalog.internal_sku == sku).first()
        if hit:
            return hit
    return (
        db.query(ProductCatalog)
        .filter(ProductCatalog.partner_id == partner_id, ProductCatalog.product_name.ilike(name.strip()))
        .first()
    )


def _upsert_product(
    db,
    *,
    row: ParsedCostRow | ParsedPriceRow,
    partner: ManufacturingPartner,
    overwrite: bool,
    summary: dict,
    sku: str | None = None,
) -> ProductCatalog | None:
    name = row.product_name
    internal_sku = sku or getattr(row, "internal_sku", None)
    if not internal_sku:
        from app.services.quotes.pricing_excel_parser import generate_internal_sku

        partner_code = getattr(row, "partner_code", None) or partner.partner_code or "OTHER"
        internal_sku = generate_internal_sku(name, partner_code=partner_code)

    existing = _find_product(db, sku=internal_sku, name=name, partner_id=partner.id)
    partner_code = getattr(row, "partner_product_code", None)
    if existing and not overwrite:
        summary["products"]["skipped"] += 1
        return existing
    if existing and overwrite:
        existing.product_name = name
        if partner_code:
            existing.partner_product_code = partner_code
        _assign_product_metadata(existing, product_name=name, partner_code=getattr(row, "partner_code", None))
        existing.notes = "excel_import"
        summary["products"]["updated"] += 1
        return existing
    category, family = _product_taxonomy(name, getattr(row, "partner_code", None) or partner.partner_code)
    product = ProductCatalog(
        partner_id=partner.id,
        internal_sku=internal_sku,
        partner_product_code=partner_code,
        product_name=name,
        product_category=category,
        product_family=family,
        status="active",
        base_currency="USD",
        default_incoterm="FOB",
        image_url=_infer_product_image_url(name, getattr(row, "partner_code", None) or partner.partner_code),
        attributes_json=_pricing_model_attributes(
            None,
            product_name=name,
            partner_code=getattr(row, "partner_code", None) or partner.partner_code,
        ),
        notes="excel_import",
    )
    db.add(product)
    db.flush()
    summary["products"]["created"] += 1
    return product


def _import_cost_row(
    db,
    row: ParsedCostRow,
    *,
    meta,
    overwrite: bool,
    summary: dict,
    partners: dict[str, ManufacturingPartner],
) -> None:
    partner = partners.get(row.partner_code) or partners.get("OTHER")
    if not partner:
        summary["errors"].append(f"no partner for {row.product_name}")
        return
    product = _upsert_product(db, row=row, partner=partner, overwrite=overwrite, summary=summary, sku=row.internal_sku)
    if not product:
        return
    existing = db.query(ProductCostModel).filter(ProductCostModel.product_id == product.id).first()
    if existing and not overwrite:
        summary["cost_models"]["skipped"] += 1
        return
    if existing and overwrite:
        cm = existing
        summary["cost_models"]["updated"] += 1
    else:
        cm = ProductCostModel(product_id=product.id, cost_currency="CNY", source="excel_import")
        db.add(cm)
        summary["cost_models"]["created"] += 1
    cm.unit_material_cost = row.material_cost
    cm.unit_weight = row.weight
    cm.domestic_transport_cost = row.domestic_transport
    cm.ocean_freight_unit_price = meta.ocean_freight_unit
    cm.domestic_profit_rate = meta.domestic_profit_rate
    cm.fob_cost_usd = row.fob_cost_usd
    cm.ddp_cost_usd = row.ddp_cost_usd
    cm.source = INTERNAL_PRICE_MODEL_SOURCE
    cm.notes = (
        "Workbook pricing basis: fixed RMB product cost plus transport cost "
        "(unit weight * ocean freight unit price), converted by USD/CNY FX. "
        "Customer quote output must show interval prices only; cost and margin remain internal."
    )


def _import_price_row(
    db,
    row: ParsedPriceRow,
    *,
    overwrite: bool,
    summary: dict,
    partners: dict[str, ManufacturingPartner],
    product_cache: dict[str, ProductCatalog],
) -> None:
    partner_code = _infer_partner_code_from_product_name(row.product_name, row.partner_code)
    partner = partners.get(partner_code) or partners.get("OTHER")
    if not partner:
        summary["warnings"].append(f"price tier skipped — no partner: {row.product_name}")
        return
    cache_key = f"{partner.id}:{row.product_name.lower()}"
    product = product_cache.get(cache_key)
    if not product:
        product = _find_product(db, sku=None, name=row.product_name, partner_id=partner.id)
    if not product:
        from app.services.quotes.pricing_excel_parser import generate_internal_sku

        sku = generate_internal_sku(row.product_name, partner_code=partner_code)
        product = _upsert_product(db, row=row, partner=partner, overwrite=overwrite, summary=summary, sku=sku)
    if not product:
        summary["warnings"].append(f"price tier skipped — product not found: {row.product_name}")
        return
    product_cache[cache_key] = product

    for inc, price, adj in (
        ("FOB", row.fob, row.fob_adjustment),
        ("DDP", row.ddp, row.ddp_adjustment),
    ):
        if price is None:
            continue
        q = db.query(ProductPriceTier).filter(
            ProductPriceTier.product_id == product.id,
            ProductPriceTier.min_qty == row.min_qty,
            ProductPriceTier.incoterm == inc,
        )
        if row.max_qty is None:
            q = q.filter(ProductPriceTier.max_qty.is_(None))
        else:
            q = q.filter(ProductPriceTier.max_qty == row.max_qty)
        tier = q.first()
        if tier and not overwrite:
            summary["price_tiers"]["skipped"] += 1
            continue
        if tier and overwrite:
            tier.final_unit_price = price
            tier.adjustment_value = adj
            tier.pricing_strategy = row.pricing_strategy
            tier.source = INTERNAL_PRICE_MODEL_SOURCE
            tier.notes = "Imported from workbook price list; customer quote displays every product interval."
            summary["price_tiers"]["updated"] += 1
        else:
            db.add(
                ProductPriceTier(
                    product_id=product.id,
                    min_qty=row.min_qty,
                    max_qty=row.max_qty,
                    incoterm=inc,
                    final_unit_price=price,
                    adjustment_value=adj,
                    currency="USD",
                    pricing_strategy=row.pricing_strategy,
                    source=INTERNAL_PRICE_MODEL_SOURCE,
                    notes="Imported from workbook price list; customer quote displays every product interval.",
                )
            )
            summary["price_tiers"]["created"] += 1


def _excel_money(value) -> Decimal | None:
    parsed = parse_decimal_cell(value)
    if parsed is None:
        return None
    return parsed.quantize(Decimal("0.01"))


def _find_quote_template_columns(ws) -> tuple[int, int, int, int, int] | None:
    for row_idx in range(1, min(ws.max_row, 80) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 12) + 1)]
        lowered = [str(value or "").strip().lower() for value in values]
        if not any(value == "products" for value in lowered):
            continue
        if not any(value == "quantity" for value in lowered):
            continue
        product_col = next(i + 1 for i, value in enumerate(lowered) if value == "products")
        qty_col = next(i + 1 for i, value in enumerate(lowered) if value == "quantity")
        fob_col = next((i + 1 for i, value in enumerate(lowered) if "fob" in value or "exw" in value), None)
        ddp_col = next((i + 1 for i, value in enumerate(lowered) if "ddp" in value), None)
        if fob_col and ddp_col:
            return row_idx, product_col, qty_col, fob_col, ddp_col
    return None


def _quote_sheet_partner_code(sheet_name: str) -> str | None:
    name = sheet_name.lower()
    if "jooboo" in name:
        return "JOOBOO"
    if "hosun" in name:
        return "HOSUN"
    return None


def _is_quote_template_stop(value) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return any(
        marker in text
        for marker in (
            "thank you",
            "terms",
            "instructions",
            "payment terms",
            "manufacturing lead time",
            "shipping information",
            "additional notes",
        )
    )


def _parse_quote_template_sheet(ws, *, partner_code: str) -> list[dict]:
    columns = _find_quote_template_columns(ws)
    if not columns:
        return []
    header_row, product_col, qty_col, fob_col, ddp_col = columns
    rows: list[dict] = []
    current_product = ""
    for row_idx in range(header_row + 1, ws.max_row + 1):
        product_raw = ws.cell(row_idx, product_col).value
        qty_raw = ws.cell(row_idx, qty_col).value
        if _is_quote_template_stop(product_raw):
            break
        product_text = str(product_raw or "").strip()
        if product_text and not product_text.startswith("="):
            current_product = re.sub(r"\s+", " ", product_text).strip()
        qty = parse_quantity_range(qty_raw)
        if not current_product or not qty:
            continue
        fob = _excel_money(ws.cell(row_idx, fob_col).value)
        ddp = _excel_money(ws.cell(row_idx, ddp_col).value)
        if fob is None and ddp is None:
            continue
        rows.append(
            {
                "partner_code": partner_code,
                "product_name": current_product,
                "min_qty": int(qty["min_qty"]),
                "max_qty": qty["max_qty"],
                "fob": fob,
                "ddp": ddp,
                "source_sheet": ws.title,
                "source_row": row_idx,
            }
        )
    return rows


def _group_quote_template_rows(wb) -> dict[tuple[str, str], list[dict]]:
    grouped: dict[tuple[str, str], list[dict]] = {}
    for sheet_name in wb.sheetnames:
        partner_code = _quote_sheet_partner_code(sheet_name)
        if not partner_code:
            continue
        ws = wb[sheet_name]
        for row in _parse_quote_template_sheet(ws, partner_code=partner_code):
            key = (row["partner_code"], row["product_name"])
            grouped.setdefault(key, []).append(row)
    return grouped


def _partner_product_code_from_name(name: str) -> str | None:
    match = re.match(r"^\s*([A-Z]{2,6}\d{3,6})\b", name.upper())
    return match.group(1) if match else None


def _hosun_cost_name_for_quote_name(name: str) -> str | None:
    text = re.sub(r"\s+", " ", name.lower())
    if text.startswith("3-leg 3-stage"):
        return "3-Leg 3-Stage Triple-Motor Rectangular Desk Frame 90x60mm"
    if "2-stage dual-motor rectangular" in text and "heavy duty" in text:
        return "2-Stage Dual-Motor Rectangular Desk Frame 90x60mm 300kg Capacity"
    if "3-stage dual-motor rectangular" in text and "3.54" in text:
        return "3-Stage Dual-Motor Rectangular Desk Frame 90x60mm"
    if "3-stage dual-motor rectangular" in text and "3.15" in text:
        return "3-Stage Dual-Motor Rectangular Desk Frame 80x50mm / 70x70mm"
    if "2-stage dual-motor rectangular" in text and "3.54" in text:
        return "2-Stage Dual-Motor Rectangular Desk Frame 90x60mm"
    if "2-stage dual-motor rectangular" in text and "3.15" in text:
        return "2-Stage Dual-Motor Rectangular Desk Frame 80x50mm / 70x70mm"
    if "2-stage dual-motor round" in text:
        return "2-Stage Dual-Motor Round Desk Frame ⌀ 70"
    if "2-stage four-motor face-to-face" in text:
        return "2-Stage Four-Motor Face-to-Face Rectangular Benching Frame 80x50mm / 70x70mm"
    if "3-stage four-motor face-to-face" in text:
        return "3-Stage Four-Motor Face-to-Face Rectangular Benching Frame 80x50mm / 70x70mm"
    return None


def _price_signature_score(db, product: ProductCatalog, quote_rows: list[dict]) -> tuple[int, int]:
    tiers = db.query(ProductPriceTier).filter(ProductPriceTier.product_id == product.id).all()
    existing: dict[tuple[int, int | None, str], Decimal] = {}
    for tier in tiers:
        if tier.final_unit_price is None:
            continue
        existing[(tier.min_qty, tier.max_qty, (tier.incoterm or "").upper())] = Decimal(str(tier.final_unit_price)).quantize(
            Decimal("0.01")
        )
    score = 0
    possible = 0
    for row in quote_rows:
        for incoterm, price in (("FOB", row.get("fob")), ("DDP", row.get("ddp"))):
            if price is None:
                continue
            possible += 1
            if existing.get((row["min_qty"], row["max_qty"], incoterm)) == price:
                score += 1
    return score, possible


def _find_product_for_quote_template(
    db,
    *,
    partner: ManufacturingPartner,
    quote_product_name: str,
    quote_rows: list[dict],
) -> ProductCatalog | None:
    exact = _find_product(db, sku=None, name=quote_product_name, partner_id=partner.id)
    if exact:
        return exact
    if (partner.partner_code or "").upper() == "HOSUN":
        canonical_name = _hosun_cost_name_for_quote_name(quote_product_name)
        if canonical_name:
            canonical = _find_product(db, sku=None, name=canonical_name, partner_id=partner.id)
            if canonical:
                return canonical
    code = _partner_product_code_from_name(quote_product_name)
    if code:
        hit = (
            db.query(ProductCatalog)
            .filter(
                ProductCatalog.partner_id == partner.id,
                (
                    ProductCatalog.partner_product_code.ilike(code)
                    | ProductCatalog.internal_sku.ilike(f"%{code}%")
                    | ProductCatalog.product_name.ilike(f"{code}%")
                ),
            )
            .first()
        )
        if hit:
            return hit
    best: tuple[int, int, ProductCatalog] | None = None
    for product in db.query(ProductCatalog).filter(ProductCatalog.partner_id == partner.id).all():
        score, possible = _price_signature_score(db, product, quote_rows)
        if possible and (best is None or score > best[0]):
            best = (score, possible, product)
    if best and best[0] >= min(4, best[1]) and Decimal(best[0]) / Decimal(best[1]) >= Decimal("0.70"):
        return best[2]
    return None


def _generate_template_sku(db, *, partner_code: str, product_name: str) -> str:
    from app.services.quotes.pricing_excel_parser import generate_internal_sku

    used = {sku for (sku,) in db.query(ProductCatalog.internal_sku).all()}
    code = _partner_product_code_from_name(product_name)
    return generate_internal_sku(product_name, partner_code=partner_code, explicit_sku=code, used=used)


def _upsert_quote_template_tier(
    db,
    *,
    product: ProductCatalog,
    row: dict,
    incoterm: str,
    price: Decimal,
    overwrite: bool,
    summary: dict,
) -> None:
    q = db.query(ProductPriceTier).filter(
        ProductPriceTier.product_id == product.id,
        ProductPriceTier.min_qty == row["min_qty"],
        ProductPriceTier.incoterm == incoterm,
    )
    if row["max_qty"] is None:
        candidates = q.all()
    else:
        candidates = q.filter(ProductPriceTier.max_qty == row["max_qty"]).all()
    tier = candidates[0] if candidates else None
    notes = (
        f"Imported from {row['source_sheet']} row {row['source_row']}; "
        "customer-visible interval quote price. Cost and margin remain internal."
    )
    if tier and not overwrite:
        summary["price_tiers"]["skipped"] += 1
        return
    if tier and overwrite:
        for duplicate in candidates[1:]:
            db.delete(duplicate)
        tier.final_unit_price = price
        tier.max_qty = row["max_qty"]
        tier.currency = "USD"
        tier.pricing_strategy = "volume"
        tier.source = CUSTOMER_QUOTE_MODEL_SOURCE
        tier.notes = notes
        summary["price_tiers"]["updated"] += 1
        return
    db.add(
        ProductPriceTier(
            product_id=product.id,
            min_qty=row["min_qty"],
            max_qty=row["max_qty"],
            incoterm=incoterm,
            final_unit_price=price,
            currency="USD",
            pricing_strategy="volume",
            source=CUSTOMER_QUOTE_MODEL_SOURCE,
            notes=notes,
        )
    )
    summary["price_tiers"]["created"] += 1


def _import_quote_template_product(
    db,
    *,
    partner: ManufacturingPartner,
    quote_product_name: str,
    quote_rows: list[dict],
    overwrite: bool,
    summary: dict,
) -> None:
    product = _find_product_for_quote_template(
        db,
        partner=partner,
        quote_product_name=quote_product_name,
        quote_rows=quote_rows,
    )
    if not product:
        sku = _generate_template_sku(db, partner_code=partner.partner_code or "OTHER", product_name=quote_product_name)
        row = ParsedPriceRow(
            partner_code=partner.partner_code,
            product_name=quote_product_name,
            min_qty=1,
            max_qty=49,
            fob=None,
            ddp=None,
            fob_adjustment=None,
            ddp_adjustment=None,
            pricing_strategy="volume",
        )
        product = _upsert_product(db, row=row, partner=partner, overwrite=overwrite, summary=summary, sku=sku)
        summary["quote_template_products"]["created"] += 1
    else:
        if overwrite:
            previous = product.product_name
            product.product_name = quote_product_name
            product.description_customer = quote_product_name
            attrs = _pricing_model_attributes(
                product.attributes_json,
                product_name=quote_product_name,
                partner_code=partner.partner_code,
            )
            if previous != quote_product_name:
                attrs.setdefault("source_cost_product_name", previous)
            attrs.update(
                {
                    "customer_quote_name": quote_product_name,
                    "quote_template_source": CUSTOMER_QUOTE_MODEL_SOURCE,
                    "quote_template_exact_interval_prices": True,
                    "quote_template_sheets": sorted({row["source_sheet"] for row in quote_rows}),
                }
            )
            product.attributes_json = attrs
            category, family = _product_taxonomy(quote_product_name, partner.partner_code)
            product.product_category = category
            product.product_family = family
            if not product.image_url:
                product.image_url = _infer_product_image_url(quote_product_name, partner.partner_code)
            product.notes = "excel_quote_template_import"
            summary["products"]["updated"] += 1
        summary["quote_template_products"]["matched"] += 1

    if not product:
        summary["warnings"].append(f"quote template product skipped: {quote_product_name}")
        return
    for row in quote_rows:
        if row["fob"] is not None:
            _upsert_quote_template_tier(
                db, product=product, row=row, incoterm="FOB", price=row["fob"], overwrite=overwrite, summary=summary
            )
        if row["ddp"] is not None:
            _upsert_quote_template_tier(
                db, product=product, row=row, incoterm="DDP", price=row["ddp"], overwrite=overwrite, summary=summary
            )


def _import_quote_template_rows(
    db,
    wb,
    *,
    overwrite: bool,
    summary: dict,
    partners: dict[str, ManufacturingPartner],
) -> int:
    grouped = _group_quote_template_rows(wb)
    for (partner_code, product_name), rows in grouped.items():
        partner = partners.get(partner_code) or partners.get("OTHER")
        if not partner:
            summary["warnings"].append(f"quote template skipped - no partner: {partner_code} {product_name}")
            continue
        _import_quote_template_product(
            db,
            partner=partner,
            quote_product_name=product_name,
            quote_rows=rows,
            overwrite=overwrite,
            summary=summary,
        )
    return sum(len(rows) for rows in grouped.values())


def _import_margin_row(db, row: ParsedMarginRow, *, overwrite: bool, summary: dict) -> None:
    q = db.query(MarginStrategyTier).filter(
        MarginStrategyTier.strategy_code == row.strategy_code,
        MarginStrategyTier.min_qty == row.min_qty,
    )
    if row.max_qty is None:
        q = q.filter(MarginStrategyTier.max_qty.is_(None))
    else:
        q = q.filter(MarginStrategyTier.max_qty == row.max_qty)
    existing = q.first()
    if existing and not overwrite:
        summary["margin_tiers"]["skipped"] += 1
        return
    if existing and overwrite:
        existing.multiplier = row.multiplier
        existing.strategy_name = row.strategy_name
        existing.notes = "excel_import"
        summary["margin_tiers"]["updated"] += 1
        return
    db.add(
        MarginStrategyTier(
            strategy_code=row.strategy_code,
            strategy_name=row.strategy_name,
            min_qty=row.min_qty,
            max_qty=row.max_qty,
            multiplier=row.multiplier,
            notes="excel_import",
        )
    )
    summary["margin_tiers"]["created"] += 1


def _import_fx(db, rate: Decimal | None, *, overwrite: bool, summary: dict) -> None:
    if rate is None:
        return
    today = date.today()
    fx_row = (
        db.query(FxRate)
        .filter(FxRate.base_currency == "USD", FxRate.quote_currency == "CNY", FxRate.rate_date == today)
        .first()
    )
    if fx_row and overwrite:
        fx_row.rate = rate
        fx_row.source = "excel_import"
        summary["fx_rates"]["updated"] += 1
    elif not fx_row:
        db.add(
            FxRate(
                base_currency="USD",
                quote_currency="CNY",
                rate=rate,
                rate_date=today,
                source="excel_import",
                is_manual_override=True,
                created_at=datetime.now(timezone.utc),
            )
        )
        summary["fx_rates"]["created"] += 1
    else:
        summary["fx_rates"]["skipped"] += 1


def _apply_sheet_report(
    db,
    report: SheetParseReport,
    *,
    overwrite: bool,
    summary: dict,
    partners: dict[str, ManufacturingPartner],
    product_cache: dict[str, ProductCatalog],
) -> None:
    if report.detected_type == "cost_model":
        if report.meta.fx_rate:
            _import_fx(db, report.meta.fx_rate, overwrite=overwrite, summary=summary)
        for row in report.cost_rows:
            _import_cost_row(db, row, meta=report.meta, overwrite=overwrite, summary=summary, partners=partners)
    elif report.detected_type == "price_tier":
        for row in report.price_rows:
            _import_price_row(
                db, row, overwrite=overwrite, summary=summary, partners=partners, product_cache=product_cache
            )
        for row in report.margin_rows:
            _import_margin_row(db, row, overwrite=overwrite, summary=summary)
    elif report.detected_type == "margin_strategy":
        for row in report.margin_rows:
            _import_margin_row(db, row, overwrite=overwrite, summary=summary)


def _backfill_inferred_images(db, partners: dict[str, ManufacturingPartner], summary: dict) -> None:
    """Backfill product images from known PartnerOS assets when workbook rows do not carry image files."""

    partner_by_id = {partner.id: code for code, partner in partners.items()}
    updated = 0
    for product in db.query(ProductCatalog).filter(ProductCatalog.image_url.is_(None)).all():
        partner_code = partner_by_id.get(product.partner_id)
        image_url = _infer_product_image_url(product.product_name or "", partner_code)
        if not image_url:
            continue
        product.image_url = image_url
        attrs = dict(product.attributes_json or {})
        attrs["image_backfill_source"] = "excel_import_name_match_to_partneros_asset"
        attrs["image_backfill_customer_safe"] = True
        product.attributes_json = attrs
        updated += 1
    summary.setdefault("image_backfill", {"updated": 0})
    summary["image_backfill"]["updated"] += updated


def run(
    file_path: Path,
    *,
    apply: bool,
    overwrite: bool,
) -> int:
    summary = {
        "partners": {"created": 0, "updated": 0, "skipped": 0},
        "products": {"created": 0, "updated": 0, "skipped": 0},
        "cost_models": {"created": 0, "updated": 0, "skipped": 0},
        "price_tiers": {"created": 0, "updated": 0, "skipped": 0},
        "margin_tiers": {"created": 0, "updated": 0, "skipped": 0},
        "fx_rates": {"created": 0, "updated": 0, "skipped": 0},
        "quote_template_products": {"created": 0, "matched": 0, "skipped": 0},
        "warnings": [],
        "errors": [],
    }
    sheet_reports: list[SheetParseReport] = []
    total_candidates = 0

    if not file_path.is_file():
        print(f"Excel file not found: {file_path}")
        print("Place workbook at local_data/报价模型与格式.xlsx (gitignored)")
        return 1

    print(f"Workbook: {file_path.resolve()}")
    wb = load_workbook(file_path, data_only=True, read_only=True)
    print(f"Available sheets: {', '.join(wb.sheetnames)}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = load_sheet_rows(ws)
        report = parse_workbook_sheet(sheet_name, rows)
        sheet_reports.append(report)
        total_candidates += report.candidate_rows
        summary["warnings"].extend(report.warnings)
        print()
        print(format_sheet_debug(report))

    quote_template_groups = _group_quote_template_rows(wb)
    quote_template_candidates = sum(len(rows) for rows in quote_template_groups.values())
    if quote_template_candidates:
        total_candidates += quote_template_candidates
        print()
        print(
            "Quote template rows: "
            f"{quote_template_candidates} tiers across {len(quote_template_groups)} products "
            "(customer-visible final interval prices)"
        )

    db = SessionLocal()
    try:
        partners: dict[str, ManufacturingPartner] = {}
        partner_codes = {"OTHER"}
        for report in sheet_reports:
            for row in report.cost_rows:
                partner_codes.add(row.partner_code)
            for row in report.price_rows:
                partner_codes.add(_infer_partner_code_from_product_name(row.product_name, row.partner_code))
        for partner_code, _ in quote_template_groups:
            partner_codes.add(partner_code)

        for code in sorted(partner_codes):
            partner, action = _ensure_partner(db, code, overwrite=overwrite)
            partners[code] = partner
            summary["partners"][action] += 1

        product_cache: dict[str, ProductCatalog] = {}
        for report in sheet_reports:
            _apply_sheet_report(
                db, report, overwrite=overwrite, summary=summary, partners=partners, product_cache=product_cache
            )
        db.flush()
        for (partner_code, product_name), rows in quote_template_groups.items():
            partner = partners.get(partner_code) or partners.get("OTHER")
            if not partner:
                summary["quote_template_products"]["skipped"] += 1
                summary["warnings"].append(f"quote template skipped - no partner: {partner_code} {product_name}")
                continue
            _import_quote_template_product(
                db,
                partner=partner,
                quote_product_name=product_name,
                quote_rows=rows,
                overwrite=overwrite,
                summary=summary,
            )
        _backfill_inferred_images(db, partners, summary)

        if apply:
            db.commit()
            print("\nImport applied.")
        else:
            db.rollback()
            print("\nDry-run — no database changes.")
    finally:
        db.close()
        wb.close()

    print("\nExcel Import Summary")
    print(f"  total candidate rows: {total_candidates}")
    for key, counts in summary.items():
        if key in ("warnings", "errors"):
            continue
        print(f"  {key}: {counts}")
    if summary["warnings"]:
        print(f"  warnings ({len(summary['warnings'])}):", summary["warnings"][:15])
    if summary["errors"]:
        print("  errors:", summary["errors"])
    return 0 if not summary["errors"] else 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Import pricing Excel into D6.2 catalog tables")
    parser.add_argument("--file", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    if args.apply and not args.confirm:
        print("Refusing --apply without --confirm")
        sys.exit(1)
    apply = args.apply and args.confirm
    sys.exit(run(Path(args.file), apply=apply, overwrite=args.overwrite))


if __name__ == "__main__":
    main()
